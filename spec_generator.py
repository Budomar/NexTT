import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os
import tempfile
import tkinter as tk
from tkinter import messagebox, ttk
import re
import pyperclip


class SpecGenerator:
    def __init__(self, main_app):
        """
        Инициализация генератора спецификаций
        
        Args:
            main_app: ссылка на главное приложение для доступа к данным
        """
        self.main_app = main_app
        self.sheets = main_app.sheets
        self.brackets_df = main_app.brackets_df
        self.entry_values = main_app.entry_values
        self.entries = main_app.entries

    def prepare_spec_data(self):
        """
        Подготавливает данные для спецификации
        Возвращает DataFrame с отсортированными данными
        """
        # Сохраняем значение из текущей активной ячейки (если есть)
        if self.main_app.root.focus_get() in self.entries.values():
            for (sheet_name, art), entry in self.entries.items():
                if entry == self.main_app.root.focus_get():
                    value = entry.get()
                    if value:
                        self.entry_values[(sheet_name, art)] = value
                    else:
                        self.entry_values.pop((sheet_name, art), None)
                    break
        
        spec_data = []
        radiator_data = []
        bracket_data = []
        brackets_temp = {}
        
        # Обработка радиаторов
        for (sheet_name, art), value in self.entry_values.items():
            if value and sheet_name in self.sheets:
                try:
                    # Получаем значение из entry_values (может быть "1+3")
                    raw_value = self.entry_values.get((sheet_name, art), "")
                    # Вычисляем сумму только при формировании спецификации
                    qty_radiator = self.parse_quantity(raw_value)
                    mask = self.sheets[sheet_name]['Артикул'] == art
                    product = self.sheets[sheet_name].loc[mask]
                    if product.empty:
                        continue
                    product = product.iloc[0]
                    radiator_type = sheet_name.split()[-1]
                    price = float(product['Цена, руб'])
                    # Получаем скидку из переменной интерфейса
                    discount = float(self.main_app.radiator_discount_var.get()) if self.main_app.radiator_discount_var.get() else 0.0
                    discounted_price = round(price * (1 - discount / 100), 2)
                    total = round(discounted_price * qty_radiator, 2)
                    # Извлекаем параметры для сортировки из наименования
                    name_parts = product['Наименование'].split('/')
                    height = int(name_parts[-2].replace('мм', '').strip())
                    length = int(name_parts[-1].replace('мм', '').strip().split()[0])
                    # Определяем Вид подключения для сортировки
                    connection_type = "VK" if "VK" in sheet_name else "K"
                    radiator_data.append({
                        "№": len(radiator_data) + 1,
                        "Артикул": str(product['Артикул']).strip(),
                        "Наименование": str(product['Наименование']),
                        "Мощность, Вт": float(product.get('Мощность, Вт', 0)),
                        "Цена, руб (с НДС)": float(price),
                        "Скидка, %": float(discount),
                        "Цена со скидкой, руб (с НДС)": float(discounted_price),
                        "Кол-во": int(qty_radiator),
                        "Сумма, руб (с НДС)": float(total),
                        "ConnectionType": connection_type,  # Для группировки VK/K
                        "RadiatorType": int(radiator_type),  # Тип радиатора (10, 11, 20 и т.д.)
                        "Height": height,  # Высота для сортировки
                        "Length": length  # Длина для сортировки
                    })
                    # Обработка кронштейнов (только если не добавлены в предпросмотре)
                    if self.main_app.bracket_var.get() != "Без кронштейнов" and not hasattr(self.main_app, 'preview_brackets_added'):
                        brackets = self.calculate_brackets(
                            radiator_type=radiator_type,
                            length=length,
                            height=height,
                            bracket_type=self.main_app.bracket_var.get(),
                            qty_radiator=qty_radiator
                        )
                        for art_bracket, qty_bracket in brackets:
                            mask_bracket = self.brackets_df['Артикул'] == art_bracket
                            bracket_info = self.brackets_df.loc[mask_bracket]
                            if bracket_info.empty:
                                continue
                            key = art_bracket.strip()
                            if key not in brackets_temp:
                                brackets_temp[key] = {
                                    "Артикул": art_bracket,
                                    "Наименование": str(bracket_info.iloc[0]['Наименование']),
                                    "Цена, руб (с НДС)": float(bracket_info.iloc[0]['Цена, руб']),
                                    "Кол-во": 0,
                                    "Сумма, руб (с НДС)": 0.0
                                }
                            price_bracket = float(bracket_info.iloc[0]['Цена, руб'])
                            # Получаем скидку на кронштейны из переменной интерфейса
                            discount_bracket = float(self.main_app.bracket_discount_var.get()) if self.main_app.bracket_discount_var.get() else 0.0
                            discounted_price_bracket = round(price_bracket * (1 - discount_bracket / 100), 2)
                            qty_total = qty_bracket
                            brackets_temp[key]["Кол-во"] += int(qty_total)
                            brackets_temp[key]["Сумма, руб (с НДС)"] += round(discounted_price_bracket * qty_total, 2)
                except Exception as e:
                    messagebox.showerror("Ошибка", f"Ошибка в данных радиатора: {str(e)}")
                    return None
        
        # Формирование данных кронштейнов
        if brackets_temp:
            for b in brackets_temp.values():
                bracket_discount = float(self.main_app.bracket_discount_var.get()) if self.main_app.bracket_discount_var.get() else 0.0
                price_with_discount = round(float(b["Цена, руб (с НДС)"]) * (1 - bracket_discount / 100), 2)
                bracket_data.append({
                    "№": len(radiator_data) + len(bracket_data) + 1,
                    "Артикул": str(b["Артикул"]),
                    "Наименование": str(b["Наименование"]),
                    "Мощность, Вт": 0.0,
                    "Цена, руб (с НДС)": float(b["Цена, руб (с НДС)"]),
                    "Скидка, %": float(bracket_discount),
                    "Цена со скидкой, руб (с НДС)": float(price_with_discount),
                    "Кол-во": int(b["Кол-во"]),
                    "Сумма, руб (с НДС)": float(b["Сумма, руб (с НДС)"]),
                    "ConnectionType": "Bracket"  # Для кронштейнов
                })
        
        # Сортировка радиаторов
        radiator_data_sorted = sorted(
            radiator_data, 
            key=lambda x: (
                0 if x["ConnectionType"] == "VK" else 1,  # Сначала VK, потом K
                x["RadiatorType"],  # Затем по типу радиатора (10, 11, 20...)
                x["Height"],  # Затем по высоте
                x["Length"]  # Затем по длине
            )
        )
        
        # Обновляем номера после сортировки
        for i, item in enumerate(radiator_data_sorted, 1):
            item["№"] = i
        
        # Объединение данных (отсортированные радиаторы + кронштейны)
        combined_data = radiator_data_sorted + bracket_data
        
        if not combined_data:
            messagebox.showwarning("Пусто", "Нет данных для формирования спецификации")
            return None
        
        # Создание DataFrame (удаляем временные поля для сортировки)
        df = pd.DataFrame(
            combined_data,
            columns=[
                "№", "Артикул", "Наименование", "Мощность, Вт",
                "Цена, руб (с НДС)", "Скидка, %",
                "Цена со скидкой, руб (с НДС)", "Кол-во",
                "Сумма, руб (с НДС)"
            ]
        )
        return df

    def save_excel_spec(self, spec_data, path, correspondence_data=None):
        """Сохраняет спецификацию в Excel с сортировкой по типоразмеру"""
        # Создаем книгу и лист
        wb = Workbook()
        ws = wb.active
        ws.title = "Спецификация"
        
        # Стили оформления
        header_font = Font(name='Calibri', size=11, bold=True)
        data_font = Font(name='Calibri', size=11)
        bold_font = Font(name='Calibri', size=11, bold=True)
        alignment_center = Alignment(horizontal='center', vertical='center')
        alignment_left = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'),
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        # Стиль для денежных значений (2 знака после запятой)
        money_style = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        
        # Заголовки столбцов
        headers = [
            "№", "Артикул", "Наименование", "Мощность, Вт",
            "Цена, руб (с НДС)", "Скидка, %", 
            "Цена со скидкой, руб (с НДС)", "Кол-во", 
            "Сумма, руб (с НДС)"
        ]
        ws.append(headers)
        
        # Применяем стили к заголовкам
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = alignment_center
            cell.border = thin_border
        
        # Разделяем радиаторы и кронштейны
        radiator_data = []
        bracket_data = []
        for _, row in spec_data.iterrows():
            if "Кронштейн" in str(row["Наименование"]):
                bracket_data.append(row)
            else:
                radiator_data.append(row)
        
        # Функция для извлечения параметров сортировки из наименования
        def get_sort_key(row):
            name = str(row["Наименование"])
            # Определяем тип подключения
            connection_type = 0 if "VK" in name else 1  # Сначала VK, потом K
            # Извлекаем тип радиатора (10, 11, 20 и т.д.)
            radiator_type = 0
            if "тип 10" in name: radiator_type = 10
            elif "тип 11" in name: radiator_type = 11
            elif "тип 20" in name: radiator_type = 20
            elif "тип 21" in name: radiator_type = 21
            elif "тип 22" in name: radiator_type = 22
            elif "тип 30" in name: radiator_type = 30
            elif "тип 33" in name: radiator_type = 33
            # Извлекаем высоту и длину
            parts = name.split('/')
            height = int(parts[-2].replace('мм', '').strip())
            length = int(parts[-1].replace('мм', '').strip().split()[0])
            return (connection_type, radiator_type, height, length)
        
        # Сортировка радиаторов
        radiator_data_sorted = sorted(radiator_data, key=get_sort_key)
        
        # Объединение данных (отсортированные радиаторы + кронштейны)
        combined_data = radiator_data_sorted + bracket_data
        
        # Заполняем Excel
        for i, row in enumerate(combined_data, 2):  # Начинаем с 2 строки
            # Для кронштейнов заменяем 0 на пустую строку в столбце мощности
            power_value = "" if "Кронштейн" in str(row["Наименование"]) else row["Мощность, Вт"]
            ws.append([
                i-1,  # №
                str(row["Артикул"]),  # Артикул как строка
                row["Наименование"],
                power_value,
                float(row['Цена, руб (с НДС)']),
                float(row['Скидка, %']),
                float(row['Цена со скидкой, руб (с НДС)']),
                int(row['Кол-во']),
                float(row['Сумма, руб (с НДС)'])
            ])
            
            # Форматируем строки данных
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=i, column=col)
                cell.font = data_font
                cell.border = thin_border
                # Устанавливаем числовые форматы
                if col in [5, 7, 9]:  # Столбцы с ценами и суммами
                    cell.number_format = money_style
                    cell.alignment = alignment_center
                elif col == 4:  # Столбец "Мощность, Вт" - центрируем
                    cell.alignment = alignment_center
                elif col in [1, 6, 8]:  # Другие числовые столбцы
                    cell.alignment = alignment_center
                else:
                    cell.alignment = alignment_left
        
        # Добавляем итоговую строку
        total_row = len(combined_data) + 2
        total_sum = spec_data["Сумма, руб (с НДС)"].sum()
        total_qty_radiators = sum(spec_data.query("Наименование.str.contains('Радиатор')")["Кол-во"].apply(self.parse_quantity))
        total_qty_brackets = sum(spec_data.query("Наименование.str.contains('Кронштейн')")["Кол-во"].apply(self.parse_quantity))
        ws.append(["Итого", "", "", "", "", "", "", f"{total_qty_radiators}/{total_qty_brackets}", total_sum])
        
        # Форматируем итоговую строку
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row, column=col)
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = alignment_center
            if col in [5, 7, 9]:  # Форматируем денежные столбцы
                cell.number_format = money_style
        
        # Добавляем вес и объем
        total_weight, total_volume = self.calculate_total_weight_and_volume(spec_data)
        
        # Пустая строка
        ws.append([])
        
        # Строка с весом
        ws.append([f"Суммарный вес радиаторов без учета упаковки и кронштейнов- {total_weight} кг."])
        ws.merge_cells(start_row=total_row + 2, start_column=1, end_row=total_row + 2, end_column=9)
        cell = ws.cell(row=total_row + 2, column=1)
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = alignment_left
        
        # Строка с объемом
        ws.append([f"Суммарный объем радиаторов без учета упаковки и кронштейнов- {total_volume} м3."])
        ws.merge_cells(start_row=total_row + 3, start_column=1, end_row=total_row + 3, end_column=9)
        cell = ws.cell(row=total_row + 3, column=1)
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = alignment_left
        
        # Настраиваем ширину столбцов для основного листа
        column_widths = {
            'A': 5,    # №
            'B': 12,   # Артикул
            'C': 60,   # Наименование
            'D': 15,   # Мощность
            'E': 20,   # Цена
            'F': 10,   # Скидка
            'G': 30,   # Цена со скидкой
            'H': 10,   # Кол-во
            'I': 20    # Сумма
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Добавляем лист с таблицей соответствия, если есть данные
        if correspondence_data is not None and not correspondence_data.empty:
            ws_corr = wb.create_sheet("Таблица соответствия")
            # Заголовки таблицы соответствия
            corr_headers = list(correspondence_data.columns)
            ws_corr.append(corr_headers)
            
            # Применяем стили к заголовкам
            for col in range(1, len(corr_headers) + 1):
                cell = ws_corr.cell(row=1, column=col)
                cell.font = header_font
                cell.alignment = alignment_center
                cell.border = thin_border
            
            # Заполняем данные
            for i, row in correspondence_data.iterrows():
                ws_corr.append(list(row))
                # Форматируем строки данных
                for col in range(1, len(corr_headers) + 1):
                    cell = ws_corr.cell(row=i+2, column=col)
                    cell.font = data_font
                    cell.border = thin_border
                    # Центрируем только столбец "Количество" (2-й столбец)
                    if col == 2:
                        cell.alignment = alignment_center
                    else:
                        cell.alignment = alignment_left
            
            # Автоподбор ширины столбцов для листа соответствия
            for col_idx, column_name in enumerate(corr_headers, 1):
                max_length = len(str(column_name))  # Начинаем с длины заголовка
                column_letter = get_column_letter(col_idx)
                # Ищем максимальную длину содержимого в столбце
                for row in ws_corr.iter_rows(min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                # Устанавливаем ширину с небольшим запасом
                adjusted_width = (max_length + 2) * 1.2
                ws_corr.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        wb.save(path)

    def generate_spec(self, file_type="excel", tree=None):
        """Генерирует итоговую спецификацию"""
        # Получаем текущие данные спецификации
        if hasattr(self.main_app, '_current_spec_data') and self.main_app._current_spec_data is not None:
            spec_data = self.main_app._current_spec_data
        else:
            # Иначе готовим данные как обычно
            spec_data = self.prepare_spec_data()
            
        if spec_data is None or spec_data.empty:
            messagebox.showwarning("Пусто", "Нет данных для сохранения")
            return
            
        if file_type == "excel":
            try:
                file_name = "Расчёт стоимости.xlsx"
                temp_dir = tempfile.gettempdir()
                
                # Генерируем уникальное имя файла, если файл уже существует
                counter = 1
                base_name = "Расчёт стоимости"
                file_path = os.path.join(temp_dir, f"{base_name}.xlsx")
                while os.path.exists(file_path):
                    file_path = os.path.join(temp_dir, f"{base_name}_{counter}.xlsx")
                    counter += 1
                    if counter > 100000000000:  # Защита от бесконечного цикла
                        raise Exception("Не удалось создать уникальное имя файла")

                # Получаем данные таблицы соответствия
                correspondence_data = None
                if hasattr(self.main_app, 'get_correspondence_data_as_df'):
                    correspondence_data = self.main_app.get_correspondence_data_as_df()

                # Сохраняем спецификацию с (возможно) таблицей соответствия
                self.save_excel_spec(spec_data, file_path, correspondence_data)

                # Попытка открыть файл с обработкой возможных ошибок
                try:
                    self.main_app.open_file_default_app(file_path)
                except Exception as open_error:
                    # Если не удалось открыть, предлагаем пользователю открыть вручную
                    if messagebox.askyesno(
                        "Ошибка открытия",
                        f"Не удалось автоматически открыть файл: {str(open_error)}\n"
                        f"Файл сохранен по пути: {file_path}\n\n"
                        "Хотите открыть его вручную?"):
                        try:
                            os.startfile(os.path.dirname(file_path))
                        except:
                            messagebox.showinfo(
                                "Информация",
                                f"Файл сохранен по пути: {file_path}\n"
                                "Откройте его вручную."
                            )
                            
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при создании Excel: {str(e)}")
                import traceback
                traceback.print_exc()
                
        elif file_type == "csv":
            file_path = tk.filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV Files", "*.csv")],
                title="Сохранить как CSV"
            )
            if not file_path:
                return
            try:
                df = pd.DataFrame({
                    "Артикул": spec_data["Артикул"],
                    "Кол-во": spec_data["Кол-во"]
                })
                df.to_csv(file_path, index=False, sep=';', encoding='utf-8-sig')
                messagebox.showinfo("Успешно", f"CSV сохранен:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка сохранения CSV:\n{str(e)}")

    def calculate_total_weight_and_volume(self, spec_data):
        """
        Рассчитывает общий вес и объем радиаторов (без учета кронштейнов)
        """
        total_weight = 0.0
        total_volume = 0.0
        # Перебираем все строки в спецификации
        for index, row in spec_data.iterrows():
            # Пропускаем строку "Итого" и кронштейны
            if row["№"] == "Итого" or "Кронштейн" in str(row["Наименование"]):
                continue
            # Получаем артикул и количество
            art = str(row["Артикул"]).strip()
            qty = int(row["Кол-во"])
            # Ищем радиатор в данных
            for sheet_name, data in self.sheets.items():
                # Проверяем наличие артикула в текущем листе
                product = data[data["Артикул"].str.strip() == art]
                if not product.empty:
                    # Суммируем вес и объем
                    total_weight += float(product.iloc[0]["Вес, кг"]) * qty
                    total_volume += float(product.iloc[0]["Объем, м3"]) * qty
                    break  # Прерываем поиск после нахождения
        # Округляем значения как в образце
        return round(total_weight, 1), round(total_volume, 3)

    def calculate_totals(self, spec_data):
        """Рассчитывает общие вес и объем (альтернативная версия)"""
        total_weight = 0.0
        total_volume = 0.0
        for _, row in spec_data.iterrows():
            art = str(row['Артикул']).strip()
            qty = int(row['Кол-во'])
            for sheet, data in self.sheets.items():
                data_arts = data['Артикул'].astype(str).str.strip()
                product = data[data_arts == art]
                if not product.empty:
                    total_weight += float(product.iloc[0]['Вес, кг']) * qty
                    total_volume += float(product.iloc[0]['Объем, м3']) * qty
                    break
        return float(total_weight), float(total_volume)

    def format_power(self, power_w):
        """
        Форматирует мощность с автоматическим выбором единиц измерения
        """
        try:
            power_w = float(power_w)
            if power_w >= 1_000_000:
                power_value = power_w / 1_000_000
                return f"{round(power_value, 3)} МВт"
            elif power_w >= 1_000:
                power_value = power_w / 1_000
                return f"{round(power_value, 3)} кВт"
            else:
                return f"{round(power_w, 2)} Вт"
        except (ValueError, TypeError):
            return f"{power_w} Вт"

    def format_weight(self, weight_kg):
        """Форматирует вес с автоматическим выбором единиц измерения"""
        if weight_kg >= 1000:
            return f"{weight_kg / 1000:.3f} т"
        else:
            return f"{weight_kg:.3f} кг"

    def parse_quantity(self, value):
        """Преобразует значение в целое число (количество радиаторов).
        Обрабатывает различные форматы: целые числа, дробные, с запятыми, с пробелами.
        Поддерживает сложение значений, разделённых знаком '+'.
        """
        if not value or pd.isna(value) or value == "":
            return 0
        str_value = str(value).strip()
        if '+' in str_value:
            parts = str_value.split('+')
            total = 0
            for part in parts:
                part = part.strip()
                if not part:
                    continue
                # Удаляем всё, кроме цифр, точек и запятых
                part_cleaned = re.sub(r'[^\d.,]', '', part)
                if not part_cleaned:
                    continue
                try:
                    # Обработка разделителей
                    if ',' in part_cleaned and '.' in part_cleaned:
                        part_cleaned = part_cleaned.replace('.', '').replace(',', '.')
                    elif ',' in part_cleaned:
                        part_cleaned = part_cleaned.replace(',', '.')
                    float_val = float(part_cleaned)
                    total += int(round(float_val))
                except (ValueError, TypeError):
                    continue
            return total
        else:
            # Обработка одиночного значения
            cleaned = re.sub(r'[^\d.,]', '', str_value)
            if not cleaned:
                return 0
            try:
                if ',' in cleaned and '.' in cleaned:
                    cleaned = cleaned.replace('.', '').replace(',', '.')
                elif ',' in cleaned:
                    cleaned = cleaned.replace(',', '.')
                float_val = float(cleaned)
                return int(round(float_val))
            except (ValueError, TypeError):
                return 0

    def calculate_brackets(self, radiator_type, length, height, bracket_type, qty_radiator=1):
        """
        Рассчитывает необходимые кронштейны для радиатора
        с учётом новой матрицы подбора (длины до 3000 мм)
        """
        brackets = []
        
        # Настенные кронштейны
        if bracket_type == "Настенные кронштейны":
            if radiator_type in ["10", "11"]:
                # Для типов 10, 11 всегда нужны левый и правый кронштейны
                brackets.extend([
                    ("К9.2L", 2 * qty_radiator),
                    ("К9.2R", 2 * qty_radiator)
                ])
                # Дополнительный средний кронштейн для длинных радиаторов
                if 1700 <= length <= 3000:
                    brackets.append(("К9.3-40", 1 * qty_radiator))
                    
            elif radiator_type in ["20", "21", "22", "30", "33"]:
                # Определяем артикул по высоте
                art_map = {
                    300: "К15.4300",
                    400: "К15.4400", 
                    500: "К15.4500",
                    600: "К15.4600",
                    900: "К15.4900"
                }
                
                if height in art_map:
                    art = art_map[height]
                    # Определяем количество по длине
                    if 400 <= length <= 1600:
                        qty = 2 * qty_radiator
                    elif 1700 <= length <= 3000:  # Изменено с 2000 на 3000
                        qty = 3 * qty_radiator
                    else:
                        qty = 0
                    
                    if qty > 0:
                        brackets.append((art, qty))
        
        # Напольные кронштейны
        elif bracket_type == "Напольные кронштейны":
            if radiator_type in ["10", "11"]:
                # Определяем основной артикул по высоте
                if 300 <= height <= 400:
                    main_art = "КНС450"
                elif 500 <= height <= 600:
                    main_art = "КНС470" 
                elif height == 900:
                    main_art = "КНС4100"
                else:
                    main_art = None
                
                if main_art:
                    # Всегда 2 основных кронштейна
                    brackets.append((main_art, 2 * qty_radiator))
                    # Дополнительный кронштейн для длинных радиаторов
                    if 1700 <= length <= 3000:  # Изменено с 2000 на 3000
                        brackets.append(("КНС430", 1 * qty_radiator))
                        
            elif radiator_type == "21":
                # Определяем артикул по высоте
                if 300 <= height <= 400:
                    art = "КНС650"
                elif 500 <= height <= 600:
                    art = "КНС670"
                elif height == 900:
                    art = "КНС6100"
                else:
                    art = None
                
                if art:
                    # Определяем количество по длине (4 диапазона)
                    if 400 <= length <= 1100:
                        qty = 2 * qty_radiator
                    elif 1200 <= length <= 1600:  # Для типа 21: 1200-1600
                        qty = 3 * qty_radiator
                    elif 1700 <= length <= 2400:  # Новый диапазон
                        qty = 4 * qty_radiator
                    elif 2500 <= length <= 3000:  # Новый диапазон
                        qty = 5 * qty_radiator
                    else:
                        qty = 0
                    
                    if qty > 0:
                        brackets.append((art, qty))
                        
            elif radiator_type in ["20", "22", "30", "33"]:
                # Определяем артикул по высоте
                if 300 <= height <= 400:
                    art = "КНС550"
                elif 500 <= height <= 600:
                    art = "КНС570"
                elif height == 900:
                    art = "КНС5100"
                else:
                    art = None
                
                if art:
                    # Определяем количество по длине (4 диапазона)
                    if 400 <= length <= 1100:
                        qty = 2 * qty_radiator
                    elif 1100 <= length <= 1600:  # Для типов 20,22,30,33: 1100-1600
                        qty = 3 * qty_radiator
                    elif 1700 <= length <= 2400:  # Новый диапазон
                        qty = 4 * qty_radiator
                    elif 2500 <= length <= 3000:  # Новый диапазон
                        qty = 5 * qty_radiator
                    else:
                        qty = 0
                    
                    if qty > 0:
                        brackets.append((art, qty))
        
        return brackets

    def calculate_total_power(self, spec_data):
        """Рассчитывает суммарную мощность (Вт) с учетом количества"""
        total_power = 0.0
        for index, row in spec_data.iterrows():
            # Пропуск итоговой строки
            if row["№"] == "Итого":
                continue
            # Получение значений из строки
            power_str = str(row["Мощность, Вт"]).strip()
            qty = row["Кол-во"]
            try:
                # Конвертация мощности в число
                power = float(power_str) if power_str not in ['', 'nan', 'None'] else 0.0
                # Расчет и суммирование
                if power >= 0 and qty >= 0:
                    total_power += power * qty
                else:
                    print(f"Некорректные значения в строке {index}: мощность={power}, количество={qty}")
            except ValueError:
                print(f"Ошибка конвертации мощности в строке {index}: '{power_str}'")
            except TypeError:
                print(f"Неправильный тип данных в строке {index}")
        return round(total_power, 2)

    def copy_column(self, spec_data, column_name):
        """Копирует данные столбца в буфер обмена, исключая итоговую строку"""
        try:
            # Фильтрация данных (исключаем строку "Итого")
            filtered_data = spec_data[spec_data["№"] != "Итого"]
            # Проверка наличия столбца
            if column_name not in filtered_data.columns:
                raise ValueError(f"Столбец {column_name} не найден")
            # Преобразование данных в строки
            data_to_copy = filtered_data[column_name].astype(str)
            # Удаление пробелов и пустых значений
            cleaned_data = data_to_copy.str.strip().replace('nan', '')
            # Копирование в буфер через переносы строк
            pyperclip.copy('\n'.join(cleaned_data))
        except KeyError as ke:
            print(f"Ошибка ключа: {str(ke)}")
        except ValueError as ve:
            print(f"Ошибка значения: {str(ve)}")
        except Exception as e:
            print(f"Общая ошибка копирования: {str(e)}")

    def copy_articul_column(self, spec_data):
        """Копирование только столбца 'Артикул' (без итоговой строки)"""
        try:
            filtered_data = spec_data[spec_data["№"] != "Итого"]
            articuls = filtered_data["Артикул"].astype(str)
            cleaned_articuls = articuls.str.strip().replace('nan', '')
            pyperclip.copy('\n'.join(cleaned_articuls))
            # Переводим окно предпросмотра на задний план и активируем главное окно
            if hasattr(self.main_app, '_preview_window') and self.main_app._preview_window.winfo_exists():
                self.main_app._preview_window.lower()
                self.main_app.root.attributes('-topmost', True)
                self.main_app.root.focus_force()
                self.main_app.root.attributes('-topmost', False)
        except Exception as e:
            print(f"Ошибка копирования артикулов: {str(e)}")

    def copy_quantity_column(self, spec_data):
        """Копирование только столбца 'Кол-во' (без итоговой строки)"""
        try:
            filtered_data = spec_data[spec_data["№"] != "Итого"]
            quantities = filtered_data["Кол-во"].astype(str)
            cleaned_quantities = quantities.str.strip().replace('nan', '')
            pyperclip.copy('\n'.join(cleaned_quantities))
            # Переводим окно предпросмотра на задний план и активируем главное окно
            if hasattr(self.main_app, '_preview_window') and self.main_app._preview_window.winfo_exists():
                self.main_app._preview_window.lower()
                self.main_app.root.attributes('-topmost', True)
                self.main_app.root.focus_force()
                self.main_app.root.attributes('-topmost', False)
        except Exception as e:
            print(f"Ошибка копирования количества: {str(e)}")

    def update_footer_labels(self, spec_data):
        """
        Возвращает словарь с текстами итоговых меток.
        Не влияет на UI напрямую.
        """
        total_power = self.calculate_total_power(spec_data)
        total_weight, total_volume = self.calculate_totals(spec_data)
        return {
            "power": self.format_power(total_power),
            "weight": self.format_weight(total_weight),
            "volume": f"{total_volume:.5f} м³"
        }


class MainApplication:
    """Главное приложение с оформлением стандартного окна Windows"""
    
    def __init__(self):
        # Создаем главное окно
        self.root = tk.Tk()
        self.setup_main_window()
        
        # Инициализируем переменные (заглушки для демонстрации)
        self.sheets = {}
        self.brackets_df = pd.DataFrame()
        self.entry_values = {}
        self.entries = {}
        self.radiator_discount_var = tk.StringVar(value="0")
        self.bracket_discount_var = tk.StringVar(value="0")
        self.bracket_var = tk.StringVar(value="Без кронштейнов")
        self._current_spec_data = None
        self._preview_window = None
        
        # Создаем экземпляр генератора спецификаций
        self.spec_generator = SpecGenerator(self)
        
        # Создаем интерфейс
        self.create_interface()
        
    def setup_main_window(self):
        """Настраивает главное окно в стиле Windows"""
        # Устанавливаем заголовок окна
        self.root.title("Генератор спецификаций радиаторов")
        
        # Устанавливаем иконку (если есть файл icon.ico в той же папке)
        try:
            self.root.iconbitmap('icon.ico')
        except:
            pass  # Если иконки нет, используем стандартную
        
        # Получаем размеры экрана
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Устанавливаем окно на весь экран, но оставляем место для панели задач
        self.root.geometry(f"{screen_width}x{screen_height-50}+0+0")
        
        # Альтернативный вариант - развернуть на весь экран
        # self.root.state('zoomed')
        
        # Минимальный размер окна
        self.root.minsize(1000, 700)
        
        # Создаем стандартное меню Windows
        self.create_menu()
        
    def create_menu(self):
        """Создает стандартное меню Windows"""
        menubar = tk.Menu(self.root)
        
        # Меню "Файл"
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Новый", command=self.new_file, accelerator="Ctrl+N")
        file_menu.add_command(label="Открыть", command=self.open_file, accelerator="Ctrl+O")
        file_menu.add_separator()
        file_menu.add_command(label="Сохранить", command=self.save_file, accelerator="Ctrl+S")
        file_menu.add_command(label="Сохранить как...", command=self.save_as_file)
        file_menu.add_separator()
        file_menu.add_command(label="Выход", command=self.exit_app, accelerator="Alt+F4")
        menubar.add_cascade(label="Файл", menu=file_menu)
        
        # Меню "Правка"
        edit_menu = tk.Menu(menubar, tearoff=0)
        edit_menu.add_command(label="Отменить", command=self.undo, accelerator="Ctrl+Z")
        edit_menu.add_command(label="Повторить", command=self.redo, accelerator="Ctrl+Y")
        edit_menu.add_separator()
        edit_menu.add_command(label="Вырезать", command=self.cut, accelerator="Ctrl+X")
        edit_menu.add_command(label="Копировать", command=self.copy, accelerator="Ctrl+C")
        edit_menu.add_command(label="Вставить", command=self.paste, accelerator="Ctrl+V")
        menubar.add_cascade(label="Правка", menu=edit_menu)
        
        # Меню "Вид"
        view_menu = tk.Menu(menubar, tearoff=0)
        view_menu.add_command(label="Полный экран", command=self.toggle_fullscreen, accelerator="F11")
        view_menu.add_command(label="Обычный размер", command=self.normal_size)
        menubar.add_cascade(label="Вид", menu=view_menu)
        
        # Меню "Справка"
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="О программе", command=self.show_about)
        menubar.add_cascade(label="Справка", menu=help_menu)
        
        self.root.config(menu=menubar)
        
        # Привязываем горячие клавиши
        self.root.bind('<F11>', lambda e: self.toggle_fullscreen())
        self.root.bind('<Escape>', lambda e: self.normal_size() if self.root.attributes('-fullscreen') else None)
        
    def create_interface(self):
        """Создает основной интерфейс приложения"""
        # Главный фрейм
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Заголовок
        title_label = ttk.Label(main_frame, text="Генератор спецификаций радиаторов", 
                               font=("Arial", 16, "bold"))
        title_label.pack(pady=(0, 20))
        
        # Область для контента
        content_frame = ttk.Frame(main_frame)
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Левая панель - ввод данных
        left_frame = ttk.LabelFrame(content_frame, text="Ввод данных", padding="10")
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        # Правая панель - результаты
        right_frame = ttk.LabelFrame(content_frame, text="Результаты", padding="10")
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        # Добавляем некоторые элементы управления для демонстрации
        ttk.Label(left_frame, text="Здесь будет интерфейс ввода данных").pack(pady=10)
        ttk.Button(left_frame, text="Добавить радиатор", command=self.add_radiator).pack(pady=5)
        ttk.Button(left_frame, text="Настройки кронштейнов", command=self.bracket_settings).pack(pady=5)
        
        ttk.Label(right_frame, text="Предпросмотр спецификации").pack(pady=10)
        ttk.Button(right_frame, text="Сгенерировать Excel", 
                  command=lambda: self.spec_generator.generate_spec("excel")).pack(pady=5)
        ttk.Button(right_frame, text="Сгенерировать CSV", 
                  command=lambda: self.spec_generator.generate_spec("csv")).pack(pady=5)
        
        # Статус бар внизу окна
        self.create_status_bar()
        
    def create_status_bar(self):
        """Создает статус бар внизу окна"""
        status_frame = ttk.Frame(self.root)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)
        
        self.status_label = ttk.Label(status_frame, text="Готов", relief=tk.SUNKEN, anchor=tk.W)
        self.status_label.pack(fill=tk.X, padx=5, pady=2)
        
    # Методы меню (заглушки)
    def new_file(self):
        self.status_label.config(text="Создание нового файла...")
        
    def open_file(self):
        self.status_label.config(text="Открытие файла...")
        
    def save_file(self):
        self.status_label.config(text="Сохранение файла...")
        
    def save_as_file(self):
        self.status_label.config(text="Сохранение файла как...")
        
    def exit_app(self):
        if messagebox.askokcancel("Выход", "Вы уверены, что хотите выйти?"):
            self.root.quit()
        
    def undo(self):
        self.status_label.config(text="Отмена действия...")
        
    def redo(self):
        self.status_label.config(text="Повтор действия...")
        
    def cut(self):
        self.status_label.config(text="Вырезание...")
        
    def copy(self):
        self.status_label.config(text="Копирование...")
        
    def paste(self):
        self.status_label.config(text="Вставка...")
        
    def toggle_fullscreen(self):
        """Переключение полноэкранного режима"""
        self.root.attributes('-fullscreen', not self.root.attributes('-fullscreen'))
        
    def normal_size(self):
        """Возврат к обычному размеру окна"""
        self.root.attributes('-fullscreen', False)
        self.root.state('normal')
        
    def show_about(self):
        """Показывает информацию о программе"""
        messagebox.showinfo("О программе", 
                          "Генератор спецификаций радиаторов\n\n"
                          "Версия 1.0\n"
                          "Программа для автоматического формирования\n"
                          "спецификаций радиаторного оборудования.")
    
    # Демонстрационные методы
    def add_radiator(self):
        messagebox.showinfo("Добавить радиатор", "Функция добавления радиатора")
        
    def bracket_settings(self):
        messagebox.showinfo("Кронштейны", "Настройки кронштейнов")
        
    def open_file_default_app(self, file_path):
        """Открывает файл в программе по умолчанию"""
        try:
            os.startfile(file_path)
        except Exception as e:
            raise e
            
    def get_correspondence_data_as_df(self):
        """Возвращает данные таблицы соответствия как DataFrame"""
        # Заглушка для демонстрации
        return pd.DataFrame()


# Запуск приложения
if __name__ == "__main__":
    app = MainApplication()
    app.root.mainloop()